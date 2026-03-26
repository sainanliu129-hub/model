#!/usr/bin/env python3
"""Embed ID/FD plots into validation xlsx sheets.

Usage:
  python3 embed_validation_figures_xlsx.py <xlsx> <id_compare> <id_error> \\
      <fd_compare> <fd_error> [<fd_q_compare> <fd_q_error>]

Optional 7th/8th args: integrated joint angle q compare/error PNG paths.
"""

import os
import sys


def _norm(p: str) -> str:
    p = (p or "").strip()
    if not p:
        return ""
    return os.path.abspath(p)


def _add_img(ws, img_path, anchor, size=(640, 360)):
    if not img_path or not os.path.isfile(img_path):
        return 0
    from openpyxl.drawing.image import Image
    img = Image(img_path)
    img.width, img.height = size
    ws.add_image(img, anchor)
    return 1


def main():
    if len(sys.argv) < 6:
        print(
            "Usage: embed_validation_figures_xlsx.py <xlsx> <id_compare> <id_error> "
            "<fd_compare> <fd_error> [<fd_q_compare> <fd_q_error>]",
            file=sys.stderr,
        )
        return 1

    xlsx = _norm(sys.argv[1])
    id_compare = _norm(sys.argv[2])
    id_error = _norm(sys.argv[3])
    fd_compare = _norm(sys.argv[4])
    fd_error = _norm(sys.argv[5])
    fd_q_compare = _norm(sys.argv[6]) if len(sys.argv) > 6 else ""
    fd_q_error = _norm(sys.argv[7]) if len(sys.argv) > 7 else ""

    try:
        from openpyxl import load_workbook
    except Exception:
        print("Need: pip install openpyxl Pillow", file=sys.stderr)
        return 2

    if not os.path.isfile(xlsx):
        print(f"xlsx not found: {xlsx}", file=sys.stderr)
        return 3

    wb = load_workbook(xlsx)
    added = 0

    if "ID_RMSE" in wb.sheetnames:
        added += _add_img(wb["ID_RMSE"], id_compare, "L2")
    if "ID_MAXERR" in wb.sheetnames:
        added += _add_img(wb["ID_MAXERR"], id_error, "L2")
    if "FD_RMSE" in wb.sheetnames:
        ws_fd = wb["FD_RMSE"]
        added += _add_img(ws_fd, fd_compare, "L2")
        if fd_q_compare:
            added += _add_img(ws_fd, fd_q_compare, "L24")
    if "FD_MAXERR" in wb.sheetnames:
        ws_fde = wb["FD_MAXERR"]
        added += _add_img(ws_fde, fd_error, "L2")
        if fd_q_error:
            added += _add_img(ws_fde, fd_q_error, "L24")

    wb.save(xlsx)
    if added <= 0:
        print("No image embedded (missing figure files or sheets).", file=sys.stderr)
        return 4
    print(f"Embedded {added} image(s).", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())


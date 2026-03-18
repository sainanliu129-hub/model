#!/usr/bin/env python3
# 在已生成的 xlsx 中嵌入附图（静摩擦/动摩擦/转动惯量）。
# 依赖: pip install openpyxl Pillow
# 用法: python3 embed_figures_xlsx.py <xlsx_path> <fig_dir> <motor_name>

import sys
import os

def main():
    if len(sys.argv) < 4:
        print("Usage: embed_figures_xlsx.py <xlsx_path> <fig_dir> <motor_name>", file=sys.stderr)
        sys.exit(1)
    xlsx_path = os.path.abspath(sys.argv[1])
    fig_dir = os.path.abspath(sys.argv[2])
    motor_name = sys.argv[3].strip()

    try:
        from openpyxl import load_workbook
        from openpyxl.drawing.image import Image
    except ImportError:
        print("Need: pip install openpyxl Pillow", file=sys.stderr)
        sys.exit(2)

    if not os.path.isfile(xlsx_path):
        print("File not found:", xlsx_path, file=sys.stderr)
        sys.exit(3)
    if not os.path.isdir(fig_dir):
        print("Figure dir not found:", fig_dir, file=sys.stderr)
        sys.exit(4)

    wb = load_workbook(xlsx_path)
    sheet_names = wb.sheetnames
    # 按名称或按顺序取 sheet（避免编码导致名称对不上）
    def get_sheet(name_or_index):
        if isinstance(name_or_index, int):
            return wb.worksheets[name_or_index] if name_or_index < len(wb.worksheets) else None
        return wb[name_or_index] if name_or_index in sheet_names else None
    idx_静摩擦 = 0 if len(sheet_names) > 0 else None
    idx_动摩擦 = 1 if len(sheet_names) > 1 else None
    idx_转动惯量 = 2 if len(sheet_names) > 2 else None

    pic_w, pic_h = 360, 270
    n_added = 0

    def add_img(ws, path, anchor="E2"):
        nonlocal n_added
        if not os.path.isfile(path):
            print("  [skip] not found:", path, file=sys.stderr)
            return
        img = Image(path)
        img.width = pic_w
        img.height = pic_h
        ws.add_image(img, anchor)
        n_added += 1

    # 静摩擦
    name = f"静摩擦辨识图_{motor_name}.png"
    path = os.path.join(fig_dir, name)
    ws1 = get_sheet("静摩擦") or (get_sheet(idx_静摩擦) if idx_静摩擦 is not None else None)
    if ws1:
        add_img(ws1, path, "E2")

    # 动摩擦
    name = f"速度拟合图_{motor_name}.png"
    path = os.path.join(fig_dir, name)
    ws2 = get_sheet("动摩擦") or (get_sheet(idx_动摩擦) if idx_动摩擦 is not None else None)
    if ws2:
        add_img(ws2, path, "E2")

    # 转动惯量：多张图纵向排列（截取后数据、惯量辨识、加速度拟合图等）
    names = [
        f"截取后数据_{motor_name}.png",
        f"惯量辨识_{motor_name}.png",
        f"加速度拟合图_{motor_name}.png",
        f"整体拟合图_{motor_name}.png",
        f"整体拟合误差图_{motor_name}.png",
        f"整体误差图_{motor_name}.png",
    ]
    ws3 = get_sheet("转动惯量") or (get_sheet(idx_转动惯量) if idx_转动惯量 is not None else None)
    if ws3:
        row = 2
        for n in names:
            path = os.path.join(fig_dir, n)
            if os.path.isfile(path):
                add_img(ws3, path, f"E{row}")
                row += 16
            else:
                print("  [skip] not found:", path, file=sys.stderr)

    wb.save(xlsx_path)

    if n_added > 0:
        print("Embedded", n_added, "image(s) into xlsx.", file=sys.stderr)
    if n_added == 0:
        print("No image embedded. Tried fig_dir:", fig_dir, "motor_name:", repr(motor_name), file=sys.stderr)
        try:
            listed = os.listdir(fig_dir)
            print("  Files in fig_dir (first 20):", listed[:20], file=sys.stderr)
        except Exception as e:
            print("  listdir error:", e, file=sys.stderr)
        sys.exit(5)
    return None

if __name__ == "__main__":
    main()
